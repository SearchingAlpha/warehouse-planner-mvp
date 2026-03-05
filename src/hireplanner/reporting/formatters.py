"""Reusable Excel styles and formatting helpers."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# Colors
BLUE = "1F4E79"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GREEN = "C6EFCE"
YELLOW = "FFEB9C"
RED = "FFC7CE"
DARK_GREEN = "006100"
DARK_YELLOW = "9C6500"
DARK_RED = "9C0006"

# Borders
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Styles
HEADER_FONT = Font(bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(bold=True, size=14, color=BLUE)
SUBTITLE_FONT = Font(bold=True, size=11, color=BLUE)

NUMBER_FORMAT_INT = "#,##0"
NUMBER_FORMAT_FLOAT = "#,##0.0"
NUMBER_FORMAT_PCT = "0.0%"
DATE_FORMAT = "YYYY-MM-DD"

# Alert fills
ALERT_FILLS = {
    "Healthy": PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid"),
    "Watch": PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid"),
    "Critical": PatternFill(start_color=RED, end_color=RED, fill_type="solid"),
}
ALERT_FONTS = {
    "Healthy": Font(color=DARK_GREEN),
    "Watch": Font(color=DARK_YELLOW),
    "Critical": Font(color=DARK_RED, bold=True),
}


def apply_header_style(ws, row: int, col_start: int, col_end: int):
    """Apply header styling to a row range."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_column_width(ws, min_width: int = 10, max_width: int = 30):
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def write_title(ws, row: int, col: int, text: str):
    """Write a title cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = TITLE_FONT
    return row + 1


def write_subtitle(ws, row: int, col: int, text: str):
    """Write a subtitle cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SUBTITLE_FONT
    return row + 1
