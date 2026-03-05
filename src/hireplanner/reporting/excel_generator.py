"""Excel report generator with 5 tabs for weekly client delivery."""
from datetime import date
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import numbers

from hireplanner.reporting.formatters import (
    apply_header_style, auto_column_width, write_title, write_subtitle,
    THIN_BORDER, NUMBER_FORMAT_INT, NUMBER_FORMAT_FLOAT, NUMBER_FORMAT_PCT,
    DATE_FORMAT, ALERT_FILLS, ALERT_FONTS, SUBTITLE_FONT,
)
from hireplanner.reporting.charts import (
    create_forecast_chart, create_backlog_chart,
    create_headcount_chart, create_accuracy_chart,
)


def _write_df_to_sheet(ws, df, start_row, start_col, locale, header_keys=None, num_formats=None):
    """Helper: write a DataFrame to a worksheet starting at (start_row, start_col).

    header_keys: list of i18n keys for column headers. If None, uses df.columns.
    num_formats: dict mapping column index (0-based) to number format string.
    """
    from hireplanner.config.i18n import t

    # Write headers
    for j, col_name in enumerate(df.columns):
        col_idx = start_col + j
        if header_keys and j < len(header_keys):
            header_text = t(header_keys[j], locale)
        else:
            header_text = col_name
        ws.cell(row=start_row, column=col_idx, value=header_text)

    apply_header_style(ws, start_row, start_col, start_col + len(df.columns) - 1)

    # Write data
    for i, (_, row) in enumerate(df.iterrows()):
        for j, val in enumerate(row):
            cell = ws.cell(row=start_row + 1 + i, column=start_col + j)
            cell.value = val
            cell.border = THIN_BORDER

            # Apply number formats
            if num_formats and j in num_formats:
                cell.number_format = num_formats[j]
            elif isinstance(val, (pd.Timestamp,)):
                cell.number_format = DATE_FORMAT
            elif isinstance(val, float):
                cell.number_format = NUMBER_FORMAT_INT

    return start_row + 1 + len(df)


def write_executive_summary(wb, config, alert_summary, accuracy_metrics, headcount_summary, locale):
    """Tab 1: Executive Summary."""
    from hireplanner.config.i18n import t

    ws = wb.create_sheet(title=t("tabs.executive_summary", locale))

    row = 1
    row = write_title(ws, row, 1, config.client_name)
    row = write_subtitle(ws, row, 1, f"{t('labels.report_date', locale)}: {date.today().isoformat()}")
    row += 1

    # Key Metrics
    row = write_subtitle(ws, row, 1, "Key Metrics")
    metrics_data = []

    if accuracy_metrics:
        metrics_data.append((t("labels.wape", locale), f"{accuracy_metrics.get('wape', 0):.1%}"))

    if alert_summary:
        metrics_data.append((
            t("labels.avg_days_of_backlog", locale),
            f"{alert_summary.get('peak_days_of_backlog', 0):.1f}"
        ))
        metrics_data.append((
            t("labels.critical_alert_days", locale),
            str(alert_summary.get("critical_days", 0))
        ))

    if headcount_summary:
        metrics_data.append((
            t("labels.peak_headcount", locale),
            str(headcount_summary.get("peak_total", 0))
        ))

    for label, value in metrics_data:
        ws.cell(row=row, column=1, value=label).font = SUBTITLE_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Alerts
    if alert_summary and alert_summary.get("critical_days", 0) > 0:
        row = write_subtitle(ws, row, 1, f"⚠ {t('alerts.critical', locale)} Alerts")
        ws.cell(row=row, column=1, value=f"{alert_summary['critical_days']} days in critical status")
        if alert_summary.get("first_critical_date"):
            ws.cell(row=row, column=2, value=f"First: {alert_summary['first_critical_date']}")
        row += 1

    row += 1

    # Trend
    trend = accuracy_metrics.get("trend", "stable") if accuracy_metrics else "stable"
    ws.cell(row=row, column=1, value=t("labels.trend", locale)).font = SUBTITLE_FONT
    ws.cell(row=row, column=2, value=t(f"labels.{trend}", locale))

    auto_column_width(ws)


def write_daily_forecast(wb, config, forecast_df, locale):
    """Tab 2: Daily Forecast (28 days) per active flow."""
    from hireplanner.config.i18n import t

    ws = wb.create_sheet(title=t("tabs.daily_forecast", locale))
    current_row = 1

    for flow in config.active_flows:
        flow_data = forecast_df[forecast_df["flow"] == flow].copy().reset_index(drop=True)
        if flow_data.empty:
            continue

        flow_label = t(f"labels.{flow}", locale)
        current_row = write_subtitle(ws, current_row, 1, flow_label)

        # Add day of week
        display_df = pd.DataFrame({
            "date": flow_data["date"],
            "day_of_week": pd.to_datetime(flow_data["date"]).dt.day_name(),
            "forecast_p50": flow_data["forecast_p50"].round(0),
            "forecast_p10": flow_data["forecast_p10"].round(0),
            "forecast_p90": flow_data["forecast_p90"].round(0),
        })

        header_keys = [
            "headers.date", "headers.day_of_week", "headers.forecast_p50",
            "headers.forecast_p10", "headers.forecast_p90",
        ]

        data_start = current_row
        current_row = _write_df_to_sheet(
            ws, display_df, current_row, 1, locale, header_keys=header_keys,
            num_formats={2: NUMBER_FORMAT_INT, 3: NUMBER_FORMAT_INT, 4: NUMBER_FORMAT_INT},
        )

        # Add chart
        chart = create_forecast_chart(
            ws, data_start, current_row - 1,
            date_col=1, p10_col=4, p50_col=3, p90_col=5,
            title=f"{flow_label} - {t('tabs.daily_forecast', locale)}",
        )
        ws.add_chart(chart, f"G{data_start}")

        current_row += 2

    auto_column_width(ws)


def write_backlog_projection(wb, config, backlog_dfs, alert_series_dict, locale):
    """Tab 3: Backlog Projection per active flow."""
    from hireplanner.config.i18n import t

    ws = wb.create_sheet(title=t("tabs.backlog_projection", locale))
    current_row = 1

    for flow in config.active_flows:
        if flow not in backlog_dfs:
            continue

        backlog_df = backlog_dfs[flow]
        alert_series = alert_series_dict.get(flow, pd.Series(["Healthy"] * len(backlog_df)))

        flow_label = t(f"labels.{flow}", locale)
        current_row = write_subtitle(ws, current_row, 1, flow_label)

        display_df = pd.DataFrame({
            "date": backlog_df["date"],
            "beg_backlog": backlog_df["beg_backlog"].round(0),
            "new_demand": backlog_df["new_demand"].round(0),
            "capacity": backlog_df["capacity"].round(0),
            "end_backlog": backlog_df["end_backlog"].round(0),
            "days_of_backlog": backlog_df["days_of_backlog"].round(2),
            "alert_status": alert_series.values,
        })

        header_keys = [
            "headers.date", "headers.beg_backlog", "headers.new_demand",
            "headers.capacity", "headers.end_backlog", "headers.days_of_backlog",
            "headers.alert_status",
        ]

        data_start = current_row
        current_row = _write_df_to_sheet(
            ws, display_df, current_row, 1, locale, header_keys=header_keys,
            num_formats={
                1: NUMBER_FORMAT_INT, 2: NUMBER_FORMAT_INT, 3: NUMBER_FORMAT_INT,
                4: NUMBER_FORMAT_INT, 5: NUMBER_FORMAT_FLOAT,
            },
        )

        # Apply alert conditional formatting
        for i in range(len(display_df)):
            cell_row = data_start + 1 + i
            status = display_df.iloc[i]["alert_status"]
            alert_cell = ws.cell(row=cell_row, column=7)
            if status in ALERT_FILLS:
                alert_cell.fill = ALERT_FILLS[status]
                alert_cell.font = ALERT_FONTS[status]

        # Backlog chart
        chart = create_backlog_chart(
            ws, data_start, current_row - 1,
            date_col=1, backlog_col=5,
            title=f"{flow_label} - {t('tabs.backlog_projection', locale)}",
        )
        ws.add_chart(chart, f"I{data_start}")

        current_row += 2

    auto_column_width(ws)


def write_headcount_plan(wb, config, headcount_df, locale):
    """Tab 4: Headcount Plan."""
    from hireplanner.config.i18n import t

    ws = wb.create_sheet(title=t("tabs.headcount_plan", locale))
    current_row = 1

    display_df = pd.DataFrame({"date": headcount_df["date"]})
    header_keys = ["headers.date"]
    hc_cols = []
    col_offset = 2

    if "outbound" in config.active_flows:
        display_df["hc_outbound"] = headcount_df["hc_outbound"]
        header_keys.append("headers.hc_outbound")
        hc_cols.append(col_offset)
        col_offset += 1

    if "inbound" in config.active_flows:
        display_df["hc_inbound"] = headcount_df["hc_inbound"]
        header_keys.append("headers.hc_inbound")
        hc_cols.append(col_offset)
        col_offset += 1

    display_df["hc_total"] = headcount_df["hc_total"]
    header_keys.append("headers.hc_total")
    hc_cols.append(col_offset)

    data_start = current_row
    num_fmts = {i: NUMBER_FORMAT_INT for i in range(1, len(display_df.columns))}
    current_row = _write_df_to_sheet(
        ws, display_df, current_row, 1, locale, header_keys=header_keys,
        num_formats=num_fmts,
    )

    # Weekly summary
    current_row += 1
    current_row = write_subtitle(ws, current_row, 1, t("labels.weekly_summary", locale))

    hc_df_copy = headcount_df.copy()
    hc_df_copy["week"] = pd.to_datetime(hc_df_copy["date"]).dt.isocalendar().week
    weekly = hc_df_copy.groupby("week").agg(
        avg_hc=("hc_total", "mean"),
        total_hours=("hc_total", lambda x: x.sum() * config.hours_per_shift),
    ).reset_index()
    weekly["avg_hc"] = weekly["avg_hc"].round(1)
    weekly["total_hours"] = weekly["total_hours"].round(0)

    weekly_headers = ["headers.date", "labels.avg_daily_hc", "labels.total_hours"]
    weekly.columns = ["Week", "Avg HC", "Total Hours"]
    _write_df_to_sheet(ws, weekly, current_row, 1, locale, header_keys=weekly_headers)

    # Chart
    chart = create_headcount_chart(
        ws, data_start, data_start + len(display_df),
        date_col=1, hc_cols=hc_cols,
        labels=[t(k, locale) for k in header_keys[1:]],
        title=t("tabs.headcount_plan", locale),
    )
    ws.add_chart(chart, f"G{data_start}")

    auto_column_width(ws)


def write_accuracy_report(wb, config, accuracy_data, locale):
    """Tab 5: Accuracy Report."""
    from hireplanner.config.i18n import t

    ws = wb.create_sheet(title=t("tabs.accuracy_report", locale))
    current_row = 1

    comparison_val = accuracy_data.get("comparison") if isinstance(accuracy_data, dict) else None
    has_comparison = comparison_val is not None and (
        not isinstance(comparison_val, pd.DataFrame) or not comparison_val.empty
    )
    if accuracy_data is None or (isinstance(accuracy_data, dict) and not has_comparison):
        # First run - no accuracy data yet
        current_row = write_title(ws, current_row, 1, t("tabs.accuracy_report", locale))
        ws.cell(row=current_row + 1, column=1, value=t("labels.no_accuracy_data", locale))
        auto_column_width(ws)
        return

    comparison_df = accuracy_data.get("comparison") if isinstance(accuracy_data, dict) else accuracy_data
    metrics = accuracy_data.get("metrics", {}) if isinstance(accuracy_data, dict) else {}
    trend = accuracy_data.get("trend", "stable") if isinstance(accuracy_data, dict) else "stable"

    # Metrics summary
    current_row = write_subtitle(ws, current_row, 1, t("tabs.accuracy_report", locale))
    current_row += 1

    for metric_name in ["wape", "mape", "mae"]:
        if metric_name in metrics:
            ws.cell(row=current_row, column=1, value=t(f"labels.{metric_name}", locale)).font = SUBTITLE_FONT
            fmt = f"{metrics[metric_name]:.1%}" if metric_name in ("wape", "mape") else f"{metrics[metric_name]:.0f}"
            ws.cell(row=current_row, column=2, value=fmt)
            current_row += 1

    ws.cell(row=current_row, column=1, value=t("labels.trend", locale)).font = SUBTITLE_FONT
    ws.cell(row=current_row, column=2, value=t(f"labels.{trend}", locale))
    current_row += 2

    # Comparison table
    if isinstance(comparison_df, pd.DataFrame) and not comparison_df.empty:
        display_df = comparison_df[["date", "forecast", "actual", "absolute_error", "percentage_error"]].copy()
        display_df["absolute_error"] = display_df["absolute_error"].round(0)
        display_df["percentage_error"] = display_df["percentage_error"].round(4)

        header_keys = [
            "headers.date", "headers.forecast_col", "headers.actual",
            "headers.error", "headers.pct_error",
        ]

        data_start = current_row
        current_row = _write_df_to_sheet(
            ws, display_df, current_row, 1, locale, header_keys=header_keys,
            num_formats={1: NUMBER_FORMAT_INT, 2: NUMBER_FORMAT_INT, 3: NUMBER_FORMAT_INT, 4: NUMBER_FORMAT_PCT},
        )

        # Accuracy chart
        chart = create_accuracy_chart(
            ws, data_start, current_row - 1,
            date_col=1, forecast_col=2, actual_col=3,
            title=t("tabs.accuracy_report", locale),
        )
        ws.add_chart(chart, f"G{data_start}")

    auto_column_width(ws)


def generate_excel_report(
    config,
    forecast_df: pd.DataFrame,
    backlog_dfs: dict[str, pd.DataFrame],
    headcount_df: pd.DataFrame,
    accuracy_data: Optional[dict] = None,
    alert_summary: Optional[dict] = None,
    alert_series_dict: Optional[dict[str, pd.Series]] = None,
    headcount_summary: Optional[dict] = None,
    output_path: str | Path = "output/report.xlsx",
    locales_dir: str | Path = "configs/locales",
) -> str:
    """Generate complete 5-tab Excel report.

    Returns path to generated file.
    """
    from hireplanner.config.i18n import load_locale

    locale = load_locale(config.language, locales_dir)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build headcount summary if not provided
    if headcount_summary is None:
        headcount_summary = {
            "peak_total": int(headcount_df["hc_total"].max()) if not headcount_df.empty else 0,
        }

    # Build accuracy metrics
    accuracy_metrics = None
    if accuracy_data and isinstance(accuracy_data, dict):
        accuracy_metrics = accuracy_data.get("metrics")
        if accuracy_metrics:
            accuracy_metrics["trend"] = accuracy_data.get("trend", "stable")

    # Tab 1: Executive Summary
    write_executive_summary(wb, config, alert_summary, accuracy_metrics, headcount_summary, locale)

    # Tab 2: Daily Forecast
    write_daily_forecast(wb, config, forecast_df, locale)

    # Tab 3: Backlog Projection
    if alert_series_dict is None:
        alert_series_dict = {}
    write_backlog_projection(wb, config, backlog_dfs, alert_series_dict, locale)

    # Tab 4: Headcount Plan
    write_headcount_plan(wb, config, headcount_df, locale)

    # Tab 5: Accuracy Report
    write_accuracy_report(wb, config, accuracy_data, locale)

    wb.save(str(output_path))
    return str(output_path)
