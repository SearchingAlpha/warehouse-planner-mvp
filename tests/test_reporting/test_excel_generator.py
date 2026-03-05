"""Tests for the Excel report generator."""
from datetime import date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from hireplanner.config.client_config import ClientConfig
from hireplanner.reporting.excel_generator import generate_excel_report
from hireplanner.reporting.formatters import apply_header_style, HEADER_FILL


# ---------------------------------------------------------------------------
# Locale YAML content (embedded so tests are self-contained)
# ---------------------------------------------------------------------------

EN_YAML = """\
tabs:
  executive_summary: "Executive Summary"
  daily_forecast: "Daily Forecast"
  backlog_projection: "Backlog Projection"
  headcount_plan: "Headcount Plan"
  accuracy_report: "Accuracy Report"

headers:
  date: "Date"
  day_of_week: "Day"
  forecast_p50: "Forecast"
  forecast_p10: "Lower (P10)"
  forecast_p90: "Upper (P90)"
  beg_backlog: "Beg Backlog"
  new_demand: "New Demand"
  capacity: "Capacity"
  end_backlog: "End Backlog"
  days_of_backlog: "Days of Backlog"
  alert_status: "Status"
  hc_inbound: "HC Inbound"
  hc_outbound: "HC Outbound"
  hc_total: "HC Total"
  forecast_col: "Forecast"
  actual: "Actual"
  error: "Abs Error"
  pct_error: "% Error"

alerts:
  healthy: "Healthy"
  watch: "Watch"
  critical: "Critical"

labels:
  client: "Client"
  report_date: "Report Date"
  forecast_period: "Forecast Period"
  wape: "WAPE"
  mape: "MAPE"
  mae: "MAE"
  avg_days_of_backlog: "Avg Days of Backlog"
  peak_headcount: "Peak Headcount"
  critical_alert_days: "Critical Alert Days"
  trend: "Week-over-Week Trend"
  improving: "Improving"
  stable: "Stable"
  degrading: "Degrading"
  no_accuracy_data: "No historical accuracy data available yet."
  outbound: "Outbound"
  inbound: "Inbound"
  total: "Total"
  weekly_summary: "Weekly Summary"
  avg_daily_hc: "Avg Daily HC"
  total_hours: "Total Hours"
"""

ES_YAML = """\
tabs:
  executive_summary: "Resumen Ejecutivo"
  daily_forecast: "Previsión Diaria"
  backlog_projection: "Proyección de Backlog"
  headcount_plan: "Plan de Personal"
  accuracy_report: "Informe de Precisión"

headers:
  date: "Fecha"
  day_of_week: "Día"
  forecast_p50: "Previsión"
  forecast_p10: "Límite Inferior (P10)"
  forecast_p90: "Límite Superior (P90)"
  beg_backlog: "Backlog Inicio"
  new_demand: "Nueva Demanda"
  capacity: "Capacidad"
  end_backlog: "Backlog Final"
  days_of_backlog: "Días de Backlog"
  alert_status: "Estado"
  hc_inbound: "Personal Entrada"
  hc_outbound: "Personal Salida"
  hc_total: "Personal Total"
  forecast_col: "Previsión"
  actual: "Real"
  error: "Error Abs"
  pct_error: "% Error"

alerts:
  healthy: "Saludable"
  watch: "Vigilancia"
  critical: "Crítico"

labels:
  client: "Cliente"
  report_date: "Fecha del Informe"
  forecast_period: "Período de Previsión"
  wape: "WAPE"
  mape: "MAPE"
  mae: "MAE"
  avg_days_of_backlog: "Promedio Días de Backlog"
  peak_headcount: "Personal Máximo"
  critical_alert_days: "Días en Alerta Crítica"
  trend: "Tendencia Semanal"
  improving: "Mejorando"
  stable: "Estable"
  degrading: "Empeorando"
  no_accuracy_data: "No hay datos históricos de precisión disponibles."
  outbound: "Salida"
  inbound: "Entrada"
  total: "Total"
  weekly_summary: "Resumen Semanal"
  avg_daily_hc: "Personal Diario Promedio"
  total_hours: "Horas Totales"
"""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_locales_dir(tmp_path, language="en"):
    """Create locale YAML files in tmp_path and return the directory path."""
    locales_dir = tmp_path / "locales"
    locales_dir.mkdir(exist_ok=True)
    (locales_dir / "en.yaml").write_text(EN_YAML, encoding="utf-8")
    (locales_dir / "es.yaml").write_text(ES_YAML, encoding="utf-8")
    return locales_dir


def _make_config(language="en", active_flows=None):
    """Build a minimal ClientConfig for testing."""
    if active_flows is None:
        active_flows = ["outbound", "inbound"]
    return ClientConfig(
        client_name="Test Warehouse",
        active_flows=active_flows,
        productivity_inbound=85.0,
        productivity_outbound=120.0,
        hours_per_shift=8,
        overhead_buffer=0.15,
        backlog_threshold_watch=1.0,
        backlog_threshold_critical=2.0,
        initial_backlog_outbound=1000,
        initial_backlog_inbound=500,
        language=language,
        forecast_horizon=28,
    )


def _make_forecast_df(n_days=7, flows=("outbound", "inbound")):
    """Build a minimal forecast DataFrame."""
    start = date(2026, 3, 1)
    rows = []
    for i in range(n_days):
        d = pd.Timestamp(start + timedelta(days=i))
        for flow in flows:
            base = 5000.0 if flow == "outbound" else 3000.0
            rows.append({
                "date": d,
                "flow": flow,
                "forecast_p50": base + i * 10,
                "forecast_p10": base * 0.8 + i * 8,
                "forecast_p90": base * 1.2 + i * 12,
            })
    return pd.DataFrame(rows)


def _make_backlog_df(n_days=7):
    """Build a minimal backlog DataFrame for a single flow."""
    start = date(2026, 3, 1)
    rows = []
    for i in range(n_days):
        d = pd.Timestamp(start + timedelta(days=i))
        rows.append({
            "date": d,
            "beg_backlog": 1000.0 + i * 50,
            "new_demand": 5000.0,
            "capacity": 4800.0,
            "end_backlog": 1200.0 + i * 50,
            "days_of_backlog": 0.25 + i * 0.01,
        })
    return pd.DataFrame(rows)


def _make_headcount_df(n_days=7):
    """Build a minimal headcount DataFrame."""
    start = date(2026, 3, 1)
    rows = []
    for i in range(n_days):
        d = pd.Timestamp(start + timedelta(days=i))
        rows.append({
            "date": d,
            "hc_outbound": 10 + i,
            "hc_inbound": 5 + i,
            "hc_total": 15 + 2 * i,
        })
    return pd.DataFrame(rows)


def _make_accuracy_data(n_days=7):
    """Build minimal accuracy data dict."""
    start = date(2026, 2, 1)
    rows = []
    for i in range(n_days):
        d = pd.Timestamp(start + timedelta(days=i))
        forecast_val = 5000.0 + i * 10
        actual_val = 5100.0 + i * 5
        rows.append({
            "date": d,
            "forecast": forecast_val,
            "actual": actual_val,
            "absolute_error": abs(forecast_val - actual_val),
            "percentage_error": abs(forecast_val - actual_val) / actual_val,
        })
    comparison_df = pd.DataFrame(rows)
    return {
        "comparison": comparison_df,
        "metrics": {"wape": 0.02, "mape": 0.021, "mae": 100.0},
        "trend": "stable",
    }


def _generate_report(tmp_path, config=None, forecast_df=None, backlog_dfs=None,
                      headcount_df=None, accuracy_data=None, alert_summary=None,
                      alert_series_dict=None):
    """Generate a report to tmp_path and return the path + loaded workbook."""
    if config is None:
        config = _make_config()
    if forecast_df is None:
        forecast_df = _make_forecast_df()
    if backlog_dfs is None:
        backlog_dfs = {
            "outbound": _make_backlog_df(),
            "inbound": _make_backlog_df(),
        }
    if headcount_df is None:
        headcount_df = _make_headcount_df()

    locales_dir = _make_locales_dir(tmp_path, config.language)
    output_path = tmp_path / "output" / "report.xlsx"

    result_path = generate_excel_report(
        config=config,
        forecast_df=forecast_df,
        backlog_dfs=backlog_dfs,
        headcount_df=headcount_df,
        accuracy_data=accuracy_data,
        alert_summary=alert_summary,
        alert_series_dict=alert_series_dict,
        output_path=output_path,
        locales_dir=locales_dir,
    )
    wb = load_workbook(result_path)
    return result_path, wb


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestExcelGenerator:
    """Tests for generate_excel_report and its tab-writing functions."""

    def test_generate_creates_file(self, tmp_path):
        """generate_excel_report creates an .xlsx file on disk."""
        result_path, _ = _generate_report(tmp_path)
        assert Path(result_path).exists()
        assert result_path.endswith(".xlsx")

    def test_workbook_has_five_tabs(self, tmp_path):
        """The generated workbook contains exactly 5 sheets."""
        _, wb = _generate_report(tmp_path)
        assert len(wb.sheetnames) == 5

    def test_tab_names_english(self, tmp_path):
        """English locale produces the correct tab names."""
        _, wb = _generate_report(tmp_path, config=_make_config(language="en"))
        expected = [
            "Executive Summary",
            "Daily Forecast",
            "Backlog Projection",
            "Headcount Plan",
            "Accuracy Report",
        ]
        assert wb.sheetnames == expected

    def test_tab_names_spanish(self, tmp_path):
        """Spanish locale produces the correct tab names."""
        _, wb = _generate_report(tmp_path, config=_make_config(language="es"))
        expected = [
            "Resumen Ejecutivo",
            "Previsión Diaria",
            "Proyección de Backlog",
            "Plan de Personal",
            "Informe de Precisión",
        ]
        assert wb.sheetnames == expected

    def test_executive_summary_has_client_name(self, tmp_path):
        """The Executive Summary tab contains the client name."""
        config = _make_config()
        _, wb = _generate_report(tmp_path, config=config)
        ws = wb["Executive Summary"]
        assert ws.cell(row=1, column=1).value == "Test Warehouse"

    def test_daily_forecast_has_data_rows(self, tmp_path):
        """Daily Forecast tab has the expected number of data rows per flow."""
        n_days = 7
        forecast_df = _make_forecast_df(n_days=n_days)
        _, wb = _generate_report(tmp_path, forecast_df=forecast_df)
        ws = wb["Daily Forecast"]

        # Count non-empty rows in column A (excluding None/empty)
        # Each flow: 1 subtitle row + 1 header row + n_days data rows + 2 spacing
        # For outbound: row 1 subtitle, row 2 header, rows 3-9 data
        # We check that row 3 has data (first data row of outbound)
        assert ws.cell(row=3, column=1).value is not None  # first data row
        assert ws.cell(row=3 + n_days - 1, column=1).value is not None  # last data row

    def test_backlog_projection_has_alert_formatting(self, tmp_path):
        """Backlog tab applies alert fill colors to status cells."""
        alert_series = pd.Series(["Healthy", "Watch", "Critical", "Healthy",
                                   "Healthy", "Watch", "Critical"])
        alert_series_dict = {
            "outbound": alert_series,
            "inbound": pd.Series(["Healthy"] * 7),
        }
        _, wb = _generate_report(tmp_path, alert_series_dict=alert_series_dict)
        ws = wb["Backlog Projection"]

        # Check that the Critical cell (row index 2+1=3 for 0-based row 2 which is
        # the third data row) in column 7 has a red fill.
        # Outbound: subtitle at row 1, header at row 2, data starts at row 3.
        # Third data row (index 2, "Critical") is at row 5.
        critical_cell = ws.cell(row=5, column=7)
        assert critical_cell.fill.start_color.rgb is not None
        # The fill should be FFC7CE (red) for Critical
        assert "FFC7CE" in str(critical_cell.fill.start_color.rgb)

    def test_headcount_plan_has_columns(self, tmp_path):
        """Headcount Plan tab header row contains expected column names."""
        _, wb = _generate_report(tmp_path)
        ws = wb["Headcount Plan"]
        # Header row is row 1 for headcount plan
        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert "Date" in headers
        assert "HC Outbound" in headers
        assert "HC Inbound" in headers
        assert "HC Total" in headers

    def test_accuracy_report_no_data(self, tmp_path):
        """Accuracy tab shows 'no data' message when accuracy_data is None."""
        _, wb = _generate_report(tmp_path, accuracy_data=None)
        ws = wb["Accuracy Report"]
        # Should have the "no data" message
        found = False
        for row in ws.iter_rows(min_row=1, max_row=5, max_col=3, values_only=True):
            for val in row:
                if val and "No historical accuracy data" in str(val):
                    found = True
                    break
        assert found, "Expected 'no data' message in Accuracy Report tab"

    def test_accuracy_report_with_data(self, tmp_path):
        """Accuracy tab writes comparison table when data is provided."""
        accuracy_data = _make_accuracy_data(n_days=5)
        _, wb = _generate_report(tmp_path, accuracy_data=accuracy_data)
        ws = wb["Accuracy Report"]

        # Look for the "Forecast" header in the comparison table
        found_forecast_header = False
        for row in ws.iter_rows(min_row=1, max_row=20, max_col=5, values_only=True):
            for val in row:
                if val == "Forecast":
                    found_forecast_header = True
                    break
        assert found_forecast_header, "Expected 'Forecast' column header in comparison table"

        # Verify data rows exist (check that numeric values are present)
        found_numeric = False
        for row in ws.iter_rows(min_row=1, max_row=20, max_col=5, values_only=True):
            for val in row:
                if isinstance(val, (int, float)) and val > 1000:
                    found_numeric = True
                    break
        assert found_numeric, "Expected numeric forecast/actual values in comparison table"

    def test_outbound_only_config(self, tmp_path):
        """Report works with outbound-only config (no inbound flow)."""
        config = _make_config(active_flows=["outbound"])
        forecast_df = _make_forecast_df(flows=("outbound",))
        backlog_dfs = {"outbound": _make_backlog_df()}
        headcount_df = _make_headcount_df()

        result_path, wb = _generate_report(
            tmp_path, config=config, forecast_df=forecast_df,
            backlog_dfs=backlog_dfs, headcount_df=headcount_df,
        )
        assert Path(result_path).exists()
        assert len(wb.sheetnames) == 5

        # Headcount tab should not have HC Inbound header
        ws = wb["Headcount Plan"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert "HC Inbound" not in headers
        assert "HC Outbound" in headers

    def test_formatters_apply_header_style(self):
        """apply_header_style applies correct font, fill, alignment, border."""
        from openpyxl import Workbook as WB
        wb = WB()
        ws = wb.active

        ws.cell(row=1, column=1, value="Col A")
        ws.cell(row=1, column=2, value="Col B")

        apply_header_style(ws, row=1, col_start=1, col_end=2)

        for col in (1, 2):
            cell = ws.cell(row=1, column=col)
            assert cell.font.bold is True
            assert cell.font.color.rgb == "00FFFFFF"  # openpyxl prepends 00
            assert cell.fill.start_color.rgb == "001F4E79"
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.wrap_text is True
            assert cell.border.left.style == "thin"
